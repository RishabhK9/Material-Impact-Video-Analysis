# -*- coding: utf-8 -*-
#!/usr/bin/env python
#Description: Extracts frames from specific video files in directory
#and outputs them to new image files. Analyzes images for ArUco marker
#and places bounding boxes around them as well as marking the coordinates
#of the midpoint. Saves all coordinates for all frames for a specific video in
#separate arrays that are outputted to plots and Excel sheets. Also combines 
#all analyzed and marked frames into a video.

# import the necessary packages
from imutils.video import VideoStream
import argparse
import imutils
import time
import cv2
import sys
import matplotlib.pyplot as plt
import numpy as np
import time
import openpyxl as xl
import os
from os.path import isfile, join
import math

__title__ = "agvTrack.py"
__author__ = "Bryant Rodriguez"
__copyright__ = "Copyright (C) 2021, Bryant Rodriguez"
__license__ = "MIT License"
__version__ = "v0.7"
__email__ = "br17d@my.fsu.edu"
__status__ = "Prototype"

xHistory = []
yHistory = []
iHistory = []
tHistory = []
degree_sign = u"\N{DEGREE SIGN}"

# Define the codec and create VideoWriter object
fourcc = cv2.VideoWriter_fourcc(*'XVID')
#out = cv2.VideoWriter('output.avi',fourcc, 20.0, (640,480))

# construct the argument parser and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-t", "--type", type=str,
	default="DICT_6X6_250",
	help="type of ArUCo tag to detect")
args = vars(ap.parse_args())

# define names of each possible ArUco tag OpenCV supports
ARUCO_DICT = {
	"DICT_4X4_50": cv2.aruco.DICT_4X4_50,
	"DICT_4X4_100": cv2.aruco.DICT_4X4_100,
	"DICT_4X4_250": cv2.aruco.DICT_4X4_250,
	"DICT_4X4_1000": cv2.aruco.DICT_4X4_1000,
	"DICT_5X5_50": cv2.aruco.DICT_5X5_50,
	"DICT_5X5_100": cv2.aruco.DICT_5X5_100,
	"DICT_5X5_250": cv2.aruco.DICT_5X5_250,
	"DICT_5X5_1000": cv2.aruco.DICT_5X5_1000,
	"DICT_6X6_50": cv2.aruco.DICT_6X6_50,
	"DICT_6X6_100": cv2.aruco.DICT_6X6_100,
	"DICT_6X6_250": cv2.aruco.DICT_6X6_250,
	"DICT_6X6_1000": cv2.aruco.DICT_6X6_1000,
	"DICT_7X7_50": cv2.aruco.DICT_7X7_50,
	"DICT_7X7_100": cv2.aruco.DICT_7X7_100,
	"DICT_7X7_250": cv2.aruco.DICT_7X7_250,
	"DICT_7X7_1000": cv2.aruco.DICT_7X7_1000,
	"DICT_ARUCO_ORIGINAL": cv2.aruco.DICT_ARUCO_ORIGINAL,
	"DICT_APRILTAG_16h5": cv2.aruco.DICT_APRILTAG_16h5,
	"DICT_APRILTAG_25h9": cv2.aruco.DICT_APRILTAG_25h9,
	"DICT_APRILTAG_36h10": cv2.aruco.DICT_APRILTAG_36h10,
	"DICT_APRILTAG_36h11": cv2.aruco.DICT_APRILTAG_36h11
}

# verify that the supplied ArUCo tag exists and is supported by
# OpenCV
if ARUCO_DICT.get(args["type"], None) is None:
	print("[INFO] ArUCo tag of '{}' is not supported".format(
		args["type"]))
	sys.exit(0)
# load the ArUCo dictionary and grab the ArUCo parameters
print("[INFO] detecting '{}' tags...".format(args["type"]))
arucoDict = cv2.aruco.Dictionary_get(ARUCO_DICT[args["type"]])
arucoParams = cv2.aruco.DetectorParameters_create()

	############################################################################
		#END OF PROGRAM INITIALIZATION. WILL NOW ANALYZE VIDEO STREAM.#
	############################################################################

# initialize the video stream and allow the camera sensor to warm up
print("[INFO] starting video stream...")
vs = VideoStream(src=0).start()
time.sleep(2.0)
i = 0
# loop over the frames from the video stream
print("Press q key to stop live video stream.")
while True:
	# grab the frame from the threed video stream and resize it
	# to have a maximum width of 1000 pixels
	frame = vs.read()
	frame = imutils.resize(frame, width=1000)
	# detect ArUco markers in the input frame
	(corners, ids, rejected) = cv2.aruco.detectMarkers(frame,
		arucoDict, parameters=arucoParams)

	# verify *at least* one ArUco marker was detected
	if len(corners) > 0:
		# flatten the ArUco IDs list
		ids = ids.flatten()
		# loop over the detected ArUCo corners
		for (markerCorner, markerID) in zip(corners, ids):
			# extract the marker corners (which are always returned
			# in top-left, top-right, bottom-right, and bottom-left
			# order)
			corners = markerCorner.reshape((4, 2))
			(topLeft, topRight, bottomRight, bottomLeft) = corners
			# convert each of the (x, y)-coordinate pairs to integers
			topRight = (int(topRight[0]), int(topRight[1]))
			bottomRight = (int(bottomRight[0]), int(bottomRight[1]))
			bottomLeft = (int(bottomLeft[0]), int(bottomLeft[1]))
			topLeft = (int(topLeft[0]), int(topLeft[1]))
			# draw the bounding box of the ArUCo detection
			cv2.line(frame, topLeft, topRight, (0, 255, 0), 2)
			cv2.line(frame, topRight, bottomRight, (0, 255, 0), 2)
			cv2.line(frame, bottomRight, bottomLeft, (0, 255, 0), 2)
			cv2.line(frame, bottomLeft, topLeft, (0, 255, 0), 2)
			
			#Calculate difference in x and y
			
			
			sY = topLeft[1] - bottomRight[1]
			sX = topLeft[0] - bottomRight[0]

			#Derive quadrant that topLeft vertex (considered the "front)
			#is pointed to.

			if (sX > 0):
				if (sY > 0):
					q = 4
				else: q = 1
			else:
				if (sY > 0):
					q = 3
				else: q = 2
				
			
			if sX != 0: #avoid "divide by zero" errors
				#Calculate slope
				#y2-y1/x2-x1

				slope1 = (sY)/(sX)
				if slope1 > 0:
					halfFlag = 1
				#Convert slope of line to radians
				rads1 = np.arctan(slope1)

				#Convert radians to degrees
				theta1 = rads1* (180/math.pi)
				theta1 = round(theta1, 2)

			#Change displayed degrees depending on active quadrant
			if q == 1:
				theta1 = abs(theta1) + 180
			elif q == 2:
				theta1 = 360 - theta1
			elif q == 3:
				theta1 = abs(theta1)
			else:
				theta1 = 180 - theta1

				

			
			# compute and draw the center (x, y)-coordinates of the
			# ArUco marker
			cX = int((topLeft[0] + bottomRight[0]) / 2.0)
			cY = int((topLeft[1] + bottomRight[1]) / 2.0)
			#Put x and y into array
			cGrid = [cX, cY]
			xHistory.append(cX)
			yHistory.append(cY)
			iHistory.append(i)
			tHistory.append(theta1)
			i = i + 1

			#Add circle to midpoint
			cv2.circle(frame, (cX, cY), 4, (0, 0, 255), -1)

			#Add circle to topLeft vertex
			cv2.circle(frame, (topLeft[0], topLeft[1]), 4, (0, 0, 255), -1)

			#Add line connecting topLeft vertex to bottomRight (the one whose slope we measured)
			cv2.line(frame, topLeft, bottomRight, (0, 0, 255, -1))

			#cv2.circle(frame, (rX, rY), 4, (0, 0, 255), -1)

			# display current coordinates of midpoint
			cv2.putText(frame, str(cGrid),
				(topLeft[0], topLeft[1] - 15),
				cv2.FONT_HERSHEY_SIMPLEX,
				0.5, (0, 255, 0), 2)

			#display current bearing of topLeft vertex (degrees)
			cv2.putText(frame, str(theta1),
				(topRight[0], topRight[1] + 15),
				cv2.FONT_HERSHEY_SIMPLEX,
				0.5, (0, 255, 0), 2)

			#display current active quadrant
			cv2.putText(frame, str(q),
				(bottomRight[0], bottomRight[1] + 15),
				cv2.FONT_HERSHEY_SIMPLEX,
				0.5, (0, 255, 0), 2)
	# show the output frame
	path = ('./liveTracking_output/')
	cv2.imwrite(os.path.join(path,'frame'+str(i)+'.jpg'), frame)
	cv2.imshow("Frame", frame)
	key = cv2.waitKey(1) & 0xFF
	# if the `q` key was pressed, break from the loop
	if key == ord("q"):
		break


	
# do a bit of cleanup
cv2.destroyAllWindows()
vs.stop()
#out.release()

	############################################################################
		#END OF VIDEO STREAM. WILL NOW CREATE PLOTS OF ALL COORDINATES.#
	############################################################################

plt.subplot(3, 1, 1)
plt.scatter(iHistory, xHistory)
plt.title("X-Coordinates Over Time")
plt.xlabel("Time (frames)")
plt.ylabel("Y-Coordinates (pixels)")


plt.subplot(3,1,2)
plt.scatter(iHistory, yHistory)
plt.title("Y-Coordinates Over Time")
plt.xlabel("Time (frames)")
plt.ylabel("X-Coordinates (pixels)")

plt.subplot(3,1,3)
plt.scatter(iHistory, tHistory)
plt.title("Theta Over Time")
plt.xlabel("Time (frames)")
plt.ylabel("Theta (degrees)")
plt.subplots_adjust(hspace=1)
#plt.show()

	############################################################################
	#  END OF PLOTTING. WILL NOW COMBINE EXTRACTED FRAMES TO CREATE NEW VIDEO. #
	############################################################################

#data.write(str(iHistory) + "\n" + str(xHistory) + "\n" + str(yHistory) + "\n")
pathIn = ('./liveTracking_output/')

iSize = len(iHistory)
finalX = xHistory[iSize-1]
finalY = yHistory[iSize-1]
finalT = tHistory[iSize-1]


fStream = open(pathIn + 'counter.txt', 'rt')
tCounter = fStream.readline()
fStream.close()
fStream = open(pathIn + 'counter.txt', 'wt')
val1 = int(tCounter)+1
fStream.write(str(val1))
fStream.write("\n" + "(" + str(finalX) + ", " + str(finalY) + ", " + str(finalT) + ")")
fStream.close()

plt.savefig("plot_" + str(val1) + '.png')

pathOut = ('./liveTracking_output/'+ 'trackedVid' + str(tCounter) + '.avi')
fps = 24
#for sorting the file names properly
frame_array = []
files = [f for f in os.listdir(pathIn) if isfile(join(pathIn, f))]

fCount = int(tCounter) + 1


	
print("Writing stream to video file...")
for i in range(2,(len(files)-fCount)):
	filename=pathIn + 'frame' + str(i) + '.jpg'
	#reading each files
	img = cv2.imread(filename)
	height, width, layers = img.shape
	size = (width,height)

	#inserting the frames into an image array
	frame_array.append(img)
out = cv2.VideoWriter(pathOut,cv2.VideoWriter_fourcc(*'DIVX'), fps, size)
for i in range(len(frame_array)):
	# writing to a image array
	out.write(frame_array[i])
out.release()

	############################################################################
		#END OF VIDEO CREATION. WILL NOW WRITE ALL DATA TO EXCEL WORKSHEET.#
	############################################################################

print("Saving data to Excel file...")
wb = xl.Workbook()

wb.save(filename='liveoutputVals' + str(val1) + '.xlsx')

iSheet = wb.create_sheet("iVals")
xSheet = wb.create_sheet("xVals")
ySheet = wb.create_sheet("yVals")
tSheet = wb.create_sheet("tVals")




for x in range(1,(iSize-1)):
	iSheet.cell(row=1, column=x, value=iHistory[x])
	xSheet.cell(row=1, column=x, value=xHistory[x])
	ySheet.cell(row=1, column=x, value=yHistory[x])
	tSheet.cell(row=1, column=x, value=tHistory[x])
wb.save('liveoutputVals'  + str(val1) + '.xlsx')


print("Editing complete. Press q to exit\n")

